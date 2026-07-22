#!/usr/bin/env python3
"""TTAG 标定启动器 — 交互式配置，自动检测续跑"""
import os, sys, glob

os.chdir(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, '.')

# ---- 固定参数 ----
DEVICE = 230030
BATH_TOLERANCE = 0.1
ADC_WINDOW = 3.0
ADC_THRESHOLD = 5
WATER_BATH_PORT = 'COM3'
TTAG_PORT = 20226

# ============================================================
print('=' * 50)
print('  TTAG 标定启动器')
print('=' * 50)

# ---- 1. 检测已有文件 ----
existing = glob.glob('calibration_data.xlsx') + glob.glob('cal_*.xlsx') + glob.glob('cal_*.csv')
resume_file = None
completed_info = None

if existing:
    latest = max(existing, key=os.path.getmtime)
    try:
        if latest.endswith('.csv'):
            import csv
            with open(latest, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                next(reader, None)
                rows = list(reader)
        else:
            from openpyxl import load_workbook
            wb = load_workbook(latest, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            wb.close()

        if rows:
            last_target = float(rows[-1][2])
            first_target = float(rows[0][2])
            step = round(float(rows[1][2]) - float(rows[0][2]), 2) if len(rows) >= 2 else 0.2
            completed_info = {
                'file': latest,
                'count': len(rows),
                'first': first_target,
                'last': last_target,
                'step': step,
            }
            print(f'\n发现已有标定文件: {latest}')
            print(f'  已完成 {len(rows)} 点: {first_target} -> {last_target} C (步长 {step})')
    except Exception as e:
        print(f'\n无法读取已有文件: {e}')

# ---- 2. 用户选择 ----
print()
if completed_info:
    print('是否从上次进度继续?')
    choice = input('[Y] 续跑  /  [N] 全新开始  (默认 Y): ').strip().lower()
    resume = choice != 'n'
else:
    print('未发现已有标定数据，将全新开始。')
    resume = False

# ---- 3. 获取参数 ----
if resume and completed_info:
    step = completed_info['step']
    start = round(completed_info['last'] + step, 1)
    print(f'\n起始温度自动设为: {start} C (上次结束于 {completed_info["last"]} C)')
    print(f'步长自动设为: {step} C (与上次一致)')

    while True:
        try:
            end_str = input(f'结束温度 (默认 50.0): ').strip()
            end = float(end_str) if end_str else 50.0
            if end <= start:
                print(f'  结束温度需大于起始温度 {start} C')
                continue
            break
        except ValueError:
            print('  请输入数字')
else:
    while True:
        try:
            s = input(f'起始温度 (默认 5.0): ').strip()
            start = float(s) if s else 5.0
            break
        except ValueError:
            print('  请输入数字')

    while True:
        try:
            e = input(f'结束温度 (默认 50.0): ').strip()
            end = float(e) if e else 50.0
            if end <= start:
                print(f'  结束温度需大于起始温度 {start} C')
                continue
            break
        except ValueError:
            print('  请输入数字')

    while True:
        try:
            s = input(f'温度步长 (默认 0.2): ').strip()
            step = float(s) if s else 0.2
            if step <= 0:
                print('  步长需大于 0')
                continue
            break
        except ValueError:
            print('  请输入数字')

# ---- 4. 预览并确认 ----
temps = []
t = start
while t <= end + step / 2:
    temps.append(round(t, 1))
    t += step

print(f'\n标定计划: {len(temps)} 点, {start} -> {end} C, 步长 {step} C')
if resume and completed_info:
    print(f'  (续跑: 已有 {completed_info["count"]} 点 + 新增 {len(temps)} 点)')

for i, tp in enumerate(temps):
    print(f'  {i+1}. {tp} C', end='  ')
    if (i + 1) % 10 == 0:
        print()
print()

confirm = input('确认开始? [Y/n]: ').strip().lower()
if confirm and confirm != 'y':
    print('已取消。')
    sys.exit(0)

# ---- 5. 启动 ----
if resume and completed_info:
    args = (f'--resume "{completed_info["file"]}"'
            f' --end {end}'
            f' --step {step}')
else:
    args = (f'--device {DEVICE}'
            f' --start {start}'
            f' --end {end}'
            f' --step {step}')

args += (f' --bath-tolerance {BATH_TOLERANCE}'
         f' --stability-window {ADC_WINDOW}'
         f' --stability-threshold {ADC_THRESHOLD}'
         f' --water-bath-port {WATER_BATH_PORT}'
         f' --ttag-port {TTAG_PORT}')

cmd = f'python -u ttag_calibration.py {args}'
print(f'\n启动: {cmd}\n')
os.system(cmd)
