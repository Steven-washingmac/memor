#!/usr/bin/env python3
"""
TTAG ADC→Temperature 自动拟合 (纯 Python)
============================================
对标 MATLAB ttag_fitting.m，标定完成后自动调用。

用法:
  python ttag_fitting.py calibration_data.xlsx     # 单独运行
  python ttag_fitting.py cal_230030_0715_1530.xlsx  # 从标定程序自动调用
"""

import os, sys
import numpy as np
import matplotlib
matplotlib.use('Agg')  # 无 GUI 后端，可在后台运行
import matplotlib.pyplot as plt

try:
    from openpyxl import load_workbook
except ImportError:
    print("需要 openpyxl: pip install openpyxl")
    sys.exit(1)


def fit_calibration(excel_path, output_dir=None):
    """
    读取标定 Excel，执行多项式拟合，生成图表和系数文件。

    参数:
        excel_path: 标定数据 Excel 文件路径
        output_dir: 输出目录 (默认与 Excel 同目录)

    返回:
        dict: {order, max_err, rmse, coeffs, adc_min, adc_max, tag_id, n_points}
    """
    if output_dir is None:
        output_dir = os.path.dirname(os.path.abspath(excel_path)) or '.'

    # ===== 1. 读取数据 =====
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # 跳过表头
    wb.close()

    tag_id = rows[0][1] if rows else '?'
    adc_raw, temp_true = [], []
    for row in rows:
        try:
            adc_v = float(row[4])      # ADC_Mean 列 (1-indexed col 5)
            temp_v = float(row[3])     # Actual(C) 列 (1-indexed col 4)
            if 0 < adc_v < 1024:
                adc_raw.append(adc_v)
                temp_true.append(temp_v)
        except (ValueError, TypeError):
            continue

    adc = np.array(adc_raw)
    temp = np.array(temp_true)
    invalid = len(rows) - len(adc)
    print(f"  读取: {len(rows)} 个数据点, 有效 {len(adc)} 个"
          + (f", 剔除 {invalid} 个异常" if invalid else ""))

    if len(adc) < 6:
        print("  ❌ 有效数据点不足，无法拟合")
        return None

    # ===== 2. 归一化 ADC 到 [-1, 1] =====
    adc_min = float(np.min(adc))
    adc_max = float(np.max(adc))
    adc_norm = 2.0 * (adc - adc_min) / (adc_max - adc_min) - 1.0

    # ===== 3. 6阶 & 7阶拟合 =====
    print(f"\n  === 多项式拟合 (6阶 vs 7阶) ===")
    print(f"  {'阶数':<6} {'MaxErr(°C)':<14} {'RMSE(°C)':<14} {'判定'}")
    print(f"  {'-'*42}")

    best = None
    for order in [6, 7]:
        try:
            coeffs = np.polyfit(adc_norm, temp, order)
            temp_pred = np.polyval(coeffs, adc_norm)
            errors = np.abs(temp - temp_pred)
            max_err = float(np.max(errors))
            rmse = float(np.sqrt(np.mean(errors ** 2)))
            ok = max_err < 0.1
            print(f"  {order:<6} {max_err:<14.4f} {rmse:<14.4f} {'✅ OK' if ok else '❌ >0.1°C'}")
            if best is None or (ok and not best['ok']):
                best = {'order': order, 'coeffs': coeffs, 'max_err': max_err,
                        'rmse': rmse, 'ok': ok, 'errors': errors}
        except Exception as e:
            print(f"  {order:<6} 拟合失败: {e}")

    if best is None:
        print("  ❌ 拟合失败")
        return None

    if best['max_err'] >= 0.1:
        print(f"\n  ⚠️ 7阶 maxErr={best['max_err']:.4f}°C > 0.1°C，精度不足！")

    # ===== 4. 绘制拟合图 =====
    base = os.path.splitext(os.path.basename(excel_path))[0]
    png_path = os.path.join(output_dir, f'{base}_fit.png')
    txt_path = os.path.join(output_dir, f'{base}_coeffs.txt')

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 5))

    # 子图1: 拟合曲线 vs 原始数据
    adc_dense = np.linspace(adc_min, adc_max, 500)
    adc_dense_norm = 2.0 * (adc_dense - adc_min) / (adc_max - adc_min) - 1.0
    temp_fit = np.polyval(best['coeffs'], adc_dense_norm)

    ax1.scatter(adc, temp, s=12, c='black', alpha=0.7, label='Measured')
    ax1.plot(adc_dense, temp_fit, 'r-', lw=2,
             label=f'{best["order"]}-order (maxErr={best["max_err"]:.4f}°C)')
    ax1.set_xlabel('ADC')
    ax1.set_ylabel('Temperature (°C)')
    ax1.set_title(f'TTAG Calibration — Tag ID={tag_id}, {len(adc)} points')
    ax1.legend(loc='best')
    ax1.grid(True, alpha=0.3)

    # 子图2: 误差分布
    ax2.scatter(adc, best['errors'], s=10, c='blue', alpha=0.6,
                label=f'{best["order"]}-order errors')
    ax2.axhline(y=0.1, color='red', ls='--', lw=1.5, label='±0.1°C threshold')
    ax2.axhline(y=0.05, color='gray', ls=':', lw=1, label='±0.05°C')
    ax2.set_xlabel('ADC')
    ax2.set_ylabel('Absolute Error (°C)')
    ax2.set_title(f'Fitting Error — {best["order"]}-order, maxErr={best["max_err"]:.4f}°C')
    ax2.legend(loc='best')
    ax2.grid(True, alpha=0.3)

    plt.tight_layout()
    fig.savefig(png_path, dpi=150, bbox_inches='tight')
    plt.close(fig)
    print(f"\n  图表: {png_path}")

    # ===== 5. 保存系数 =====
    with open(txt_path, 'w', encoding='utf-8') as f:
        f.write(f"TTAG ADC→Temperature 拟合系数\n")
        f.write(f"标签ID: {tag_id}\n")
        f.write(f"数据点: {len(adc)}\n")
        f.write(f"拟合阶数: {best['order']}\n")
        f.write(f"最大误差: {best['max_err']:.4f}°C\n")
        f.write(f"RMSE: {best['rmse']:.4f}°C\n")
        f.write(f"ADC 范围: {round(adc_min)} ~ {round(adc_max)}\n")
        f.write(f"多项式系数 (归一化 ADC → [-1,1]):\n")
        for i, c in enumerate(best['coeffs']):
            f.write(f"  p{i} = {c:.15e}\n")
    print(f"  系数: {txt_path}")

    # ===== 6. 打印 Python 可用代码 =====
    print(f"\n  === Python adc_to_temperature() ===")
    print(f"  ADC_MIN = {round(adc_min)}")
    print(f"  ADC_MAX = {round(adc_max)}")
    print(f"  COEFFS = [")
    coeff_str = ', '.join(f'{c:.12e}' for c in best['coeffs'])
    print(f"      {coeff_str}")
    print(f"  ]")
    print(f"")
    print(f"  def adc_to_temperature(adc):")
    print(f"      if adc <= 0 or adc >= 1024:")
    print(f"          return None")
    print(f"      x = 2.0 * (adc - ADC_MIN) / (ADC_MAX - ADC_MIN) - 1.0")
    print(f"      return np.polyval(COEFFS, x)")

    return {
        'order': best['order'],
        'max_err': best['max_err'],
        'rmse': best['rmse'],
        'coeffs': best['coeffs'],
        'adc_min': adc_min,
        'adc_max': adc_max,
        'tag_id': tag_id,
        'n_points': len(adc),
        'png_path': png_path,
        'txt_path': txt_path,
    }


def main():
    if len(sys.argv) < 2:
        print("用法: python ttag_fitting.py <calibration_data.xlsx>")
        print("      python ttag_fitting.py cal_230030_0715_1530.xlsx")
        sys.exit(1)

    excel_path = sys.argv[1]
    if not os.path.exists(excel_path):
        print(f"❌ 文件不存在: {excel_path}")
        sys.exit(1)

    print(f"TTAG ADC→Temperature 自动拟合")
    print(f"文件: {excel_path}\n")

    result = fit_calibration(excel_path)

    if result:
        print(f"\n✅ 拟合完成: {result['order']}阶, maxErr={result['max_err']:.4f}°C")
    else:
        print(f"\n❌ 拟合失败")
        sys.exit(1)


if __name__ == '__main__':
    main()
